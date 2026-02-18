"""
routes/statistics.py — Global statistics page.

Provides:
- GET /statistics          — Global statistics overview for all gardens
- GET /statistics/print    — Print-friendly view
- GET /statistics/excel    — Excel export of statistics

Shows aggregated statistics across all gardens including:
- Total number of gardens and their names
- Total number of beds for all gardens
- Total number of sub-beds (active, reserve, total)
- Number of sub-beds per crop in each category
"""

from flask import Blueprint, render_template, send_file
from database import get_gardens, get_garden_stats, get_crops, get_categories, get_cycles, get_db
from datetime import date

statistics_bp = Blueprint('statistics', __name__, url_prefix='/statistics')


def get_global_statistics():
    """Compute global statistics across all gardens.

    Returns a dict with:
    - total_gardens: number of gardens
    - gardens: list of garden dicts with name, code
    - total_beds: total beds across all gardens
    - total_sub_beds: total sub-beds across all gardens
    - active_sub_beds: total active sub-beds
    - reserve_sub_beds: total reserve sub-beds
    - crops_by_category: dict of category -> list of {crop_name, count}
    """
    gardens = get_gardens()
    categories = get_categories()

    stats = {
        'total_gardens': len(gardens),
        'gardens': [],
        'total_beds': 0,
        'total_sub_beds': 0,
        'active_sub_beds': 0,
        'reserve_sub_beds': 0,
        'crops_by_category': {cat: [] for cat in categories},
    }

    # Aggregate garden stats
    for garden in gardens:
        garden_stats = get_garden_stats(garden['id'])
        if garden_stats:
            stats['gardens'].append({
                'id': garden['id'],
                'code': garden['garden_code'],
                'name': garden['name'],
                'beds': garden_stats['beds'],
                'total_sub_beds': garden_stats['total_sub_beds'],
                'active_sub_beds': garden_stats['active_sub_beds'],
                'reserve_sub_beds': garden_stats['reserve_sub_beds'],
            })
            stats['total_beds'] += garden_stats['beds']
            stats['total_sub_beds'] += garden_stats['total_sub_beds']
            stats['active_sub_beds'] += garden_stats['active_sub_beds']
            stats['reserve_sub_beds'] += garden_stats['reserve_sub_beds']

    # Count crops by category across all gardens (from latest cycles)
    conn = get_db()
    try:
        # Get crop counts from cycle_plans across all gardens
        # Using actual values if present, otherwise planned values
        crop_counts = conn.execute("""
            SELECT
                c.category,
                c.crop_name,
                c.id as crop_id,
                COUNT(*) as count
            FROM cycle_plans cp
            JOIN crops c ON COALESCE(cp.actual_crop_id, cp.planned_crop_id) = c.id
            JOIN sub_beds sb ON cp.sub_bed_id = sb.id
            WHERE sb.is_reserve = 0
              AND cp.cycle = (
                  SELECT MAX(cycle) FROM cycle_plans cp2 WHERE cp2.garden_id = cp.garden_id
              )
            GROUP BY c.category, c.crop_name
            ORDER BY c.category, count DESC, c.crop_name
        """).fetchall()

        for row in crop_counts:
            cat = row['category']
            if cat in stats['crops_by_category']:
                stats['crops_by_category'][cat].append({
                    'crop_name': row['crop_name'],
                    'count': row['count'],
                })
    finally:
        conn.close()

    return stats


@statistics_bp.route('/')
def index():
    """Global statistics page."""
    stats = get_global_statistics()
    categories = get_categories()

    return render_template('statistics.html',
                          stats=stats,
                          categories=categories)


@statistics_bp.route('/print')
def print_view():
    """Print-friendly version of global statistics."""
    stats = get_global_statistics()
    categories = get_categories()
    today = date.today().strftime('%d/%m/%Y')

    return render_template('statistics_print.html',
                          stats=stats,
                          categories=categories,
                          print_date=today)


@statistics_bp.route('/excel')
def export_excel():
    """Export global statistics as Excel file."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from utils.backup import backup_db

    # Auto-backup before export
    backup_db('statistics_export')

    stats = get_global_statistics()
    categories = get_categories()

    # Category colors
    CATEGORY_FILLS = {
        'Feuille': PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid'),
        'Graine': PatternFill(start_color='FFB300', end_color='FFB300', fill_type='solid'),
        'Racine': PatternFill(start_color='00897B', end_color='00897B', fill_type='solid'),
        'Fruit': PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid'),
        'Couverture': PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid'),
    }

    HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    HEADER_FILL = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    TITLE_FONT = Font(name='Calibri', bold=True, size=14)
    CELL_BORDER = Border(
        bottom=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
    )

    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Résumé"

    # Title
    ws_summary.cell(row=1, column=1, value="Statistiques Globales").font = TITLE_FONT
    ws_summary.merge_cells('A1:D1')

    # Date
    ws_summary.cell(row=2, column=1, value=f"Date: {date.today().strftime('%d/%m/%Y')}")

    # Summary stats
    row = 4
    ws_summary.cell(row=row, column=1, value="Nombre de jardins:")
    ws_summary.cell(row=row, column=2, value=stats['total_gardens'])
    row += 1
    ws_summary.cell(row=row, column=1, value="Total planches:")
    ws_summary.cell(row=row, column=2, value=stats['total_beds'])
    row += 1
    ws_summary.cell(row=row, column=1, value="Total sous-planches:")
    ws_summary.cell(row=row, column=2, value=stats['total_sub_beds'])
    row += 1
    ws_summary.cell(row=row, column=1, value="Sous-planches actives:")
    ws_summary.cell(row=row, column=2, value=stats['active_sub_beds'])
    row += 1
    ws_summary.cell(row=row, column=1, value="Sous-planches en réserve:")
    ws_summary.cell(row=row, column=2, value=stats['reserve_sub_beds'])

    # Gardens list
    row += 2
    ws_summary.cell(row=row, column=1, value="Liste des jardins:").font = Font(bold=True)
    row += 1
    headers = ['Code', 'Nom', 'Planches', 'Sous-planches', 'Actives', 'Réserve']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    row += 1

    for garden in stats['gardens']:
        ws_summary.cell(row=row, column=1, value=garden['code'])
        ws_summary.cell(row=row, column=2, value=garden['name'])
        ws_summary.cell(row=row, column=3, value=garden['beds'])
        ws_summary.cell(row=row, column=4, value=garden['total_sub_beds'])
        ws_summary.cell(row=row, column=5, value=garden['active_sub_beds'])
        ws_summary.cell(row=row, column=6, value=garden['reserve_sub_beds'])
        row += 1

    # Column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 12

    # Sheet 2: Crops by category
    ws_crops = wb.create_sheet(title="Cultures par catégorie")

    ws_crops.cell(row=1, column=1, value="Cultures par catégorie (cycles en cours)").font = TITLE_FONT
    ws_crops.merge_cells('A1:C1')

    row = 3
    for category in categories:
        crops = stats['crops_by_category'].get(category, [])

        # Category header
        cat_cell = ws_crops.cell(row=row, column=1, value=category)
        cat_cell.font = HEADER_FONT
        if category in CATEGORY_FILLS:
            cat_cell.fill = CATEGORY_FILLS[category]
        ws_crops.merge_cells(f'A{row}:C{row}')
        row += 1

        # Column headers
        ws_crops.cell(row=row, column=1, value="Culture").font = Font(bold=True)
        ws_crops.cell(row=row, column=2, value="Sous-planches").font = Font(bold=True)
        row += 1

        if crops:
            for crop in crops:
                ws_crops.cell(row=row, column=1, value=crop['crop_name'])
                ws_crops.cell(row=row, column=2, value=crop['count'])
                row += 1
        else:
            ws_crops.cell(row=row, column=1, value="(aucune culture)")
            row += 1

        row += 1  # Empty row between categories

    ws_crops.column_dimensions['A'].width = 25
    ws_crops.column_dimensions['B'].width = 15
    ws_crops.column_dimensions['C'].width = 15

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"statistiques_globales_{date.today().strftime('%Y%m%d')}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
