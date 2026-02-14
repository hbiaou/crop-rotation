"""
utils/export.py — Excel export generation using pandas + openpyxl.

Generates .xlsx files with one sheet per garden, styled header row.
Columns: Planche, Sous-planche, Catégorie, Culture, Notes.

See FEATURES_SPEC.md section F9.
"""

from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import get_map_data, get_gardens, get_garden


# Category colors for header/cell styling (from AGENTS.md)
CATEGORY_FILLS = {
    'Feuille': PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid'),
    'Graine': PatternFill(start_color='FFB300', end_color='FFB300', fill_type='solid'),
    'Racine': PatternFill(start_color='00897B', end_color='00897B', fill_type='solid'),
    'Fruit': PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid'),
    'Couverture': PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid'),
}

HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style='thin', color='0D47A1'),
    right=Side(style='thin', color='E2E8F0'),
)
CELL_BORDER = Border(
    bottom=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
)


def _build_sheet(ws, map_data):
    """Populate a worksheet with map data rows and styled header."""
    # Header row
    columns = ['Planche', 'Sous-planche', 'Catégorie', 'Culture', 'Notes']
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    # Data rows — sorted by bed_number, then sub-bed position
    row_idx = 2
    for bed in map_data['beds']:
        for sb in bed['sub_beds']:
            # Use actual data if available, otherwise planned
            category = sb.get('actual_category') or sb.get('planned_category', '')
            crop = sb.get('actual_crop_name') or sb.get('planned_crop_name', '')

            ws.cell(row=row_idx, column=1, value=f"P{bed['bed_number']:02d}").border = CELL_BORDER
            ws.cell(row=row_idx, column=2, value=f"S{sb['position']}").border = CELL_BORDER

            cat_cell = ws.cell(row=row_idx, column=3, value=category)
            cat_cell.border = CELL_BORDER
            # Apply category color fill
            if category in CATEGORY_FILLS:
                cat_cell.fill = CATEGORY_FILLS[category]
                cat_cell.font = Font(color='FFFFFF', bold=True)

            ws.cell(row=row_idx, column=4, value=crop or '').border = CELL_BORDER
            ws.cell(row=row_idx, column=5, value=sb.get('notes') or '').border = CELL_BORDER
            row_idx += 1

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30

    # Freeze header row
    ws.freeze_panes = 'A2'


def generate_excel(garden_id, cycle):
    """Generate an Excel workbook for a single garden+cycle.

    Returns:
        (BytesIO buffer, filename) on success, (None, None) on failure.
    """
    import openpyxl

    garden = get_garden(garden_id)
    if not garden:
        return None, None

    map_data = get_map_data(garden_id, cycle)
    if not map_data or not map_data['beds']:
        return None, None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = garden['garden_code']

    _build_sheet(ws, map_data)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"rotation_{garden['garden_code']}_{cycle}.xlsx"
    return buffer, filename


def generate_excel_all(cycle):
    """Generate an Excel workbook with one sheet per garden for a cycle.

    Returns:
        (BytesIO buffer, filename) on success, (None, None) on failure.
    """
    import openpyxl

    gardens = get_gardens()
    if not gardens:
        return None, None

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheets_created = 0
    for garden in gardens:
        map_data = get_map_data(garden['id'], cycle)
        if not map_data or not map_data['beds']:
            continue

        ws = wb.create_sheet(title=garden['garden_code'])
        _build_sheet(ws, map_data)
        sheets_created += 1

    if sheets_created == 0:
        return None, None

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"rotation_tous_{cycle}.xlsx"
    return buffer, filename
